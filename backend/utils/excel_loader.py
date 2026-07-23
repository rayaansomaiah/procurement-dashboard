"""
Loader for the Replenishment Indent master sheet.

The sheet layout:
  Row 1  — global inputs (Machine count in M1, Monthly usage hrs in O1)
  Row 2  — column headers
  Row 3+ — one row per SKU

Per-SKU applicability % (what fraction of the machine fleet uses this SKU)
is stored only inside the MDP/CDP cell formula, in one of these shapes:
    =$M$1*0.75     -> 0.75
    =500*0.4       -> 0.40      (500 = hardcoded machine base)
    =500           -> 1.00      (no multiplier = 100%)
    =$M$1          -> 1.00
We parse the multiplier out (openpyxl, data_only=False) so it survives even
when the UI machine count differs from the sheet's hardcoded base.
"""

import io
import re
import openpyxl

# Header text (stripped + lowercased) -> internal column name
_HEADER_ALIASES = {
    "sku": "sku_code", "sku code": "sku_code", "part no": "sku_code",
    "category": "category",
    "sub category": "sub_category", "sub cat": "sub_category",
    "item": "item", "item name": "item", "description": "item",
    "brand": "brand",
    "qoh": "qoh_excel", "quantity on hand": "qoh_excel",
    "purchase price": "purchase_price", "price": "purchase_price",
    "previous 5m sales": "prev_sales_excel", "previous sales": "prev_sales_excel",
    "wallet qty": "wallet_qty",
    "mdp/cdp": "mdp_cell", "mdp": "mdp_cell",
    "consumption hrs": "consumption_hrs", "consumption hours": "consumption_hrs",
    "flf": "flf",
    "arc (weeks)": "arc_excel", "arc": "arc_excel",
}

_MACHINE_REF = re.compile(r"^\$?[A-Z]+\$?\d+$")  # e.g. $M$1, M1


def _parse_applicability(formula, machine_base: float) -> float:
    """Extract the applicability fraction from an MDP/CDP cell formula."""
    if not isinstance(formula, str):
        return 1.0
    s = formula.lstrip("=").replace(" ", "")
    if not s:
        return 1.0
    tokens = s.split("*")

    mults: list[float] = []
    for tok in tokens:
        if _MACHINE_REF.match(tok):
            continue  # a cell reference like $M$1 -> the machine base
        try:
            mults.append(float(tok))
        except ValueError:
            continue  # unrecognised token — ignore

    # Drop one numeric token equal to the machine base (the hardcoded 500).
    if machine_base and machine_base in mults:
        mults.remove(machine_base)

    if not mults:
        return 1.0
    result = 1.0
    for m in mults:
        result *= m
    return result


def _num(val, default=0.0) -> float:
    try:
        if val is None or val == "":
            return default
        return float(val)
    except (ValueError, TypeError):
        return default


def _str(val) -> str:
    return "" if val is None else str(val).strip()


def _find_header_row(ws) -> int:
    """Return the 1-based row index whose first cells contain 'SKU'."""
    for r in range(1, min(11, ws.max_row + 1)):
        for c in range(1, min(6, ws.max_column + 1)):
            if _str(ws.cell(row=r, column=c).value).lower() == "sku":
                return r
    return 2  # sensible default for this sheet


def load_indent_excel(file) -> tuple[list[dict], list[str]]:
    """
    Parse the indent master sheet.
    Returns (rows, warnings) where rows is a list of per-SKU dicts.
    """
    warnings: list[str] = []
    if isinstance(file, (bytes, bytearray)):
        file = io.BytesIO(file)

    try:
        # NOTE: read_only=False is intentional. We do random cell access
        # (ws.cell(row, col)) throughout; openpyxl's read-only mode re-seeks
        # from the start on each random access, which is catastrophically slow
        # on a BytesIO stream (the real upload path). Normal mode loads the
        # sheet into memory — fast random access, fine for a few hundred rows.
        wb = openpyxl.load_workbook(file, data_only=False, read_only=False)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")

    ws = wb[wb.sheetnames[0]]
    if len(wb.sheetnames) > 1:
        warnings.append(f"Multiple sheets found. Using first sheet: '{ws.title}'.")

    header_row = _find_header_row(ws)

    # Machine base value (M1) — used only to identify the base token in formulas.
    machine_base = _num(ws.cell(row=1, column=13).value, 0.0)  # column M = 13

    # Map each column index -> internal name
    col_map: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        key = _str(ws.cell(row=header_row, column=c).value).lower()
        if key in _HEADER_ALIASES and _HEADER_ALIASES[key] not in col_map.values():
            col_map[c] = _HEADER_ALIASES[key]

    if "sku_code" not in col_map.values():
        raise ValueError("Could not find a 'SKU' column in the uploaded sheet.")
    if "mdp_cell" not in col_map.values():
        warnings.append("No 'MDP/CDP' column found — applicability defaults to 100%.")

    rows: list[dict] = []
    missing_sku = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw: dict = {}
        for c, name in col_map.items():
            raw[name] = ws.cell(row=r, column=c).value

        sku = _str(raw.get("sku_code"))
        item = _str(raw.get("item"))
        # Skip fully-empty rows
        if not sku and not item:
            continue
        if not sku:
            sku = item  # fall back to item name as identifier
            missing_sku += 1

        applicability = _parse_applicability(raw.get("mdp_cell"), machine_base)

        rows.append({
            "sku_code":        sku,
            "category":        _str(raw.get("category")),
            "sub_category":    _str(raw.get("sub_category")),
            "item":            item or sku,
            "brand":           _str(raw.get("brand")),
            "purchase_price":  _num(raw.get("purchase_price")),
            "wallet_qty":      _num(raw.get("wallet_qty"), 0.0),
            "consumption_hrs": _num(raw.get("consumption_hrs"), 0.0),
            "flf":             _num(raw.get("flf"), 0.0),
            "applicability":   applicability,
            "prev_sales_excel": _num(raw.get("prev_sales_excel"), 0.0),
        })

    wb.close()

    if not rows:
        raise ValueError("No SKU rows found in the uploaded sheet.")
    if missing_sku:
        warnings.append(f"{missing_sku} row(s) have no SKU — tracked by Item name instead.")

    return rows, warnings
